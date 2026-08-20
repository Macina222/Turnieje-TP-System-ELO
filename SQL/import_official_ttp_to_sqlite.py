#!/usr/bin/env python3
"""
Import oficjalnych danych TTP z pliku XLSX do znormalizowanej bazy SQLite.

Przykład użycia:
    python import_official_ttp_to_sqlite.py "_Oficjalne dane.xlsx" ttp_official.sqlite

Założenia dla aktualnego pliku:
- nagłówek znajduje się w pierwszym niepustym wierszu,
- wymagane kolumny: dancers id, pair id, season, turnament code, turnament name,
  cat code, pair, group, rank, points before, points, medals, points after, medals after,
- dancers id ma postać np. "591-1411",
- pair ma postać np. "PIOTR HERTRAMPF, WERONIKA BUKOWSKA".
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from migrations import ensure_schema, CURRENT_SCHEMA_VERSION

REQUIRED_COLUMNS = {
    "dancers id",
    "pair id",
    "season",
    "turnament code",
    "turnament name",
    "cat code",
    "pair",
    "group",
    "rank",
    "points before",
    "points",
    "medals",
    "points after",
    "medals after",
}

ROMAN_CATEGORIES = ["VIII", "VII", "VI", "V", "IV", "III", "II", "I"]


@dataclass(frozen=True)
class OfficialRow:
    row_number: int
    raw_dancers_id: str
    pair_id: int
    season: int
    tournament_code: str
    tournament_name: str
    cat_code: str
    pair_name: str
    group_name: str | None
    rank: float
    points_before: float | None
    points_awarded: float | None
    medals_awarded: float | None
    points_after: float | None
    medals_after: float | None
    event_date: str | None = None


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def normalize_key(value: Any) -> str | None:
    text = normalize_text(value)
    return text.upper() if text else None


def as_int(value: Any, field_name: str, row_number: int) -> int:
    if value is None or value == "":
        raise ValueError(f"Brak wartości w kolumnie {field_name!r}, wiersz {row_number}")
    return int(float(value))


def as_float_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def parse_category(cat_code: str) -> tuple[str | None, str | None]:
    code = normalize_key(cat_code)
    if not code:
        return None, None
    for base in ROMAN_CATEGORIES:
        if code.startswith(base):
            suffix = code[len(base):]
            return base, suffix or None
    return None, None


def split_dancer_ids(raw_dancers_id: str) -> list[int]:
    parts = [part.strip() for part in str(raw_dancers_id).split("-") if part.strip()]
    return [int(part) for part in parts]


def split_pair_names(pair_name: str) -> list[str]:
    return [normalize_text(part) or "" for part in str(pair_name).split(",") if normalize_text(part)]


def find_header(rows: Iterable[tuple[int, tuple[Any, ...]]]) -> tuple[int, dict[str, int]]:
    for row_number, row in rows:
        normalized = {normalize_key(cell): idx for idx, cell in enumerate(row) if normalize_key(cell)}
        lowercase_map = {key.lower(): idx for key, idx in normalized.items()}
        if REQUIRED_COLUMNS.issubset(lowercase_map.keys()):
            return row_number, lowercase_map
    raise RuntimeError("Nie znaleziono wiersza nagłówka z wymaganymi kolumnami.")


def _extract_event_date(columns: dict[str, int]) -> str | None:
    """
    Extract event date from column mapping.

    The official XLSX doesn't have a date column yet, but future versions
    might add one. This function is a placeholder that returns None for now
    but is ready to parse a date column when it's added to the schema.
    """
    # Future: check for "data" or "date" column
    # if "data" in columns: return parse_and_normalize_date(...)

    # Currently no date in official format
    return None


def iter_official_rows(xlsx_path: Path, sheet_name: str | None = None) -> Iterable[OfficialRow]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    all_rows = list(enumerate(worksheet.iter_rows(values_only=True), start=1))
    header_row_number, columns = find_header(all_rows)

    for row_number, row in all_rows:
        if row_number <= header_row_number:
            continue
        if not any(cell is not None for cell in row):
            continue

        def get(column_name: str) -> Any:
            idx = columns[column_name]
            return row[idx] if idx < len(row) else None

        # W oficjalnym arkuszu mogą wystąpić techniczne/puste rekordy rankingowe
        # (np. lokata bez przypisanej pary). Nie są one wynikiem pary, więc importer
        # pomija je zamiast przerywać cały import.
        if not normalize_text(get("dancers id")) or get("pair id") in (None, "") or not normalize_text(get("pair")):
            continue

        yield OfficialRow(
            row_number=row_number,
            raw_dancers_id=normalize_text(get("dancers id")) or "",
            pair_id=as_int(get("pair id"), "pair id", row_number),
            season=as_int(get("season"), "season", row_number),
            tournament_code=normalize_key(get("turnament code")) or "",
            tournament_name=normalize_text(get("turnament name")) or "",
            cat_code=normalize_key(get("cat code")) or "",
            pair_name=normalize_text(get("pair")) or "",
            group_name=normalize_text(get("group")),
            rank=as_float_or_none(get("rank")) or 0.0,
            points_before=as_float_or_none(get("points before")),
            points_awarded=as_float_or_none(get("points")),
            medals_awarded=as_float_or_none(get("medals")),
            points_after=as_float_or_none(get("points after")),
            medals_after=as_float_or_none(get("medals after")),
            event_date=_extract_event_date(columns),
        )


def upsert_source_file(conn: sqlite3.Connection, xlsx_path: Path) -> int:
    conn.execute(
        """
        INSERT OR IGNORE INTO source_files(file_name, absolute_path)
        VALUES (?, ?)
        """,
        (xlsx_path.name, str(xlsx_path.resolve())),
    )
    return conn.execute(
        "SELECT source_file_id FROM source_files WHERE file_name = ? AND absolute_path = ?",
        (xlsx_path.name, str(xlsx_path.resolve())),
    ).fetchone()[0]


def get_or_create_group(conn: sqlite3.Connection, group_name: str | None) -> int | None:
    if not group_name:
        return None
    conn.execute(
        "INSERT OR IGNORE INTO groups(group_name, normalized_name) VALUES (?, ?)",
        (group_name, normalize_key(group_name)),
    )
    return conn.execute("SELECT group_id FROM groups WHERE group_name = ?", (group_name,)).fetchone()[0]


def get_or_create_tournament(conn: sqlite3.Connection, row: OfficialRow) -> int:
    conn.execute(
        """
        INSERT OR IGNORE INTO tournaments(season, tournament_code, tournament_name, event_date)
        VALUES (?, ?, ?, ?)
        """,
        (row.season, row.tournament_code, row.tournament_name, row.event_date),
    )
    return conn.execute(
        "SELECT tournament_id FROM tournaments WHERE season = ? AND tournament_code = ?",
        (row.season, row.tournament_code),
    ).fetchone()[0]


def get_or_create_event(conn: sqlite3.Connection, tournament_id: int, cat_code: str) -> int:
    base_category, class_code = parse_category(cat_code)
    conn.execute(
        """
        INSERT OR IGNORE INTO events(tournament_id, cat_code, base_category, class_code)
        VALUES (?, ?, ?, ?)
        """,
        (tournament_id, cat_code, base_category, class_code),
    )
    return conn.execute(
        "SELECT event_id FROM events WHERE tournament_id = ? AND cat_code = ?",
        (tournament_id, cat_code),
    ).fetchone()[0]


def upsert_pair_and_dancers(conn: sqlite3.Connection, row: OfficialRow, source_file_id: int) -> None:
    dancer_ids = split_dancer_ids(row.raw_dancers_id)
    dancer_names = split_pair_names(row.pair_name)

    conn.execute(
        """
        INSERT INTO pairs(pair_id, display_name, normalized_name)
        VALUES (?, ?, ?)
        ON CONFLICT(pair_id) DO UPDATE SET
            display_name = excluded.display_name,
            normalized_name = excluded.normalized_name
        """,
        (row.pair_id, row.pair_name, normalize_key(row.pair_name)),
    )

    if len(dancer_ids) != len(dancer_names):
        conn.execute(
            """
            INSERT INTO import_warnings(source_file_id, source_row_number, warning_type, message)
            VALUES (?, ?, ?, ?)
            """,
            (
                source_file_id,
                row.row_number,
                "DANCER_COUNT_MISMATCH",
                f"Liczba ID tancerzy ({len(dancer_ids)}) różni się od liczby nazwisk ({len(dancer_names)}): {row.raw_dancers_id} / {row.pair_name}",
            ),
        )

    for idx, dancer_id in enumerate(dancer_ids, start=1):
        full_name = dancer_names[idx - 1] if idx - 1 < len(dancer_names) else None
        conn.execute(
            """
            INSERT INTO dancers(dancer_id, full_name, normalized_name)
            VALUES (?, ?, ?)
            ON CONFLICT(dancer_id) DO UPDATE SET
                full_name = COALESCE(excluded.full_name, dancers.full_name),
                normalized_name = COALESCE(excluded.normalized_name, dancers.normalized_name)
            """,
            (dancer_id, full_name, normalize_key(full_name)),
        )
        conn.execute(
            """
            INSERT OR IGNORE INTO pair_members(pair_id, dancer_id, member_order)
            VALUES (?, ?, ?)
            """,
            (row.pair_id, dancer_id, idx),
        )


def import_xlsx_to_sqlite(xlsx_path: Path, sqlite_path: Path, sheet_name: str | None = None, replace: bool = False) -> None:
    if replace and sqlite_path.exists():
        sqlite_path.unlink()

    # Ensure schema is up to date (runs migrations if needed)
    ensure_schema(sqlite_path, CURRENT_SCHEMA_VERSION)

    with sqlite3.connect(sqlite_path) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        source_file_id = upsert_source_file(conn, xlsx_path)

        imported = 0
        for row in iter_official_rows(xlsx_path, sheet_name=sheet_name):
            upsert_pair_and_dancers(conn, row, source_file_id)
            group_id = get_or_create_group(conn, row.group_name)
            tournament_id = get_or_create_tournament(conn, row)
            event_id = get_or_create_event(conn, tournament_id, row.cat_code)

            conn.execute(
                """
                INSERT INTO results(
                    event_id, pair_id, group_id, rank,
                    points_before, points_awarded, medals_awarded,
                    points_after, medals_after,
                    source_file_id, source_row_number,
                    raw_dancers_id, raw_pair_name
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    event_id,
                    row.pair_id,
                    group_id,
                    row.rank,
                    row.points_before,
                    row.points_awarded,
                    row.medals_awarded,
                    row.points_after,
                    row.medals_after,
                    source_file_id,
                    row.row_number,
                    row.raw_dancers_id,
                    row.pair_name,
                ),
            )
            imported += 1

        conn.commit()

        summary = conn.execute(
            """
            SELECT
                (SELECT COUNT(*) FROM dancers) AS dancers,
                (SELECT COUNT(*) FROM pairs) AS pairs,
                (SELECT COUNT(*) FROM tournaments) AS tournaments,
                (SELECT COUNT(*) FROM events) AS events,
                (SELECT COUNT(*) FROM results) AS results,
                (SELECT COUNT(*) FROM import_warnings) AS warnings
            """
        ).fetchone()

    print(f"Zaimportowano rekordów z arkusza: {imported}")
    print(f"SQLite zapisano w: {sqlite_path}")
    print(
        "Podsumowanie bazy: "
        f"tancerze={summary[0]}, pary={summary[1]}, turnieje={summary[2]}, "
        f"eventy={summary[3]}, wyniki={summary[4]}, ostrzeżenia={summary[5]}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import oficjalnych danych TTP XLSX do SQLite.")
    parser.add_argument("xlsx", type=Path, help="Ścieżka do pliku XLSX z oficjalnymi danymi TTP.")
    parser.add_argument("sqlite", type=Path, help="Ścieżka wyjściowa do pliku SQLite.")
    parser.add_argument("--sheet", default=None, help="Nazwa arkusza. Domyślnie aktywny arkusz.")
    parser.add_argument("--replace", action="store_true", help="Usuń istniejącą bazę przed importem.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    import_xlsx_to_sqlite(args.xlsx, args.sqlite, sheet_name=args.sheet, replace=args.replace)


if __name__ == "__main__":
    main()
